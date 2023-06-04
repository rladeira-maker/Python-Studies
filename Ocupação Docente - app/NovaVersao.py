#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      sn73093
#
# Created:     28/10/2021
# Copyright:   (c) sn73093 2021
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import openpyxl

class Aula:
    #Propriedades
    def __init__(self,docente='',horario='',unidade_curricular='',indice=0):
        self.docente =docente
        self.horario=horario
        self.unidade_curricular=unidade_curricular
        self.indice=indice


class SubTurma:
    def __init__(self,aulas=[],periodo='',numero_de_alunos='16',sub_turma='A'):
        self.aulas = aulas
        self.periodo=periodo
        #self.nome=nome
        self.numero_de_alunos=numero_de_alunos
        self.sub_turma = sub_turma

class Turma(SubTurma):
    def __init__(self,nome='A'):
        self.nome = nome
        pass
    #Propriedades
##    def __init__(self,aulas,periodo,nome,numero_de_alunos):
##        self.aulas =aulas
##        self.periodo=periodo
##        self.nome=nome
##        self.numero_de_alunos=numero_de_alunos

def main():
    wb=openpyxl.load_workbook('teste.xlsx')
    planilha_ativa = wb.sheetnames[0]
    planilha = wb[planilha_ativa]
    # turma1=Turma()
    # subturmaA=SubTurma()
    # subturmaB=SubTurma()
    # subturmaA.aulas = Aula()
    aula_m1a=Aula()
    aulas_m1a=[]
    for a in range(50):
        aulas_m1a.append(aula_m1a)
    print(aulas_m1a[aulas_m1a.docente]    

    # count = 0
    # for x in range(10,2):
    #     for y in range(6):
    #         aulas_m1a.append(sheet.cell(row=y+1,column=y+1).value
    #         count += 1
    #     pass
    # print(((subturmaA.aulas[1])))

if __name__ == '__main__':
    main()
