#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      SN73093
#
# Created:     29/06/2021
# Copyright:   (c) SN73093 2021
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import openpyxl
from os import system, name

class Turma:
    '''A turma terá aulas, ambientes e docentes'''


    def turma(coord1,sheet1,qtd_aulas):
        aulas = {}
        diasDaSemana = ('seg1','seg2','ter1','ter2','qua1','qua2','qui1','qui2','sex1','sex2')
        if qtd_aulas == 4:
            aulasDoDia = ('A1','A2','A3','A4','doc1','doc2','amb1','amb2')
        else:
            aulasDoDia = ('A1','A2','A3','A4','A5','A6','doc1','doc2','amb1','amb2')

        docentesDoDia = ('doc1,doc2,doc3,doc4')
        ambientesDoDia = ('amb1,amb2,amb3,amb4')
        x = coord1[0]
        y = coord1[1]
        y += 2
        title = sheet1.cell(column=1,row=y).value
        for dia in range(len(diasDaSemana)):

            for aula in range(len(aulasDoDia)):
                aulas[diasDaSemana[dia],aulasDoDia[aula]] = sheet1.cell(column=x+dia,row=y+aula).value
                if aula == 7 and '2' in str(diasDaSemana[dia]):
                    x += 1

        return (title,aulas)

def clearScreen():

    # for windows os
    if name == 'nt':
        _ = system('cls')

    # for mac and linux os(The name is posix)
    else:
        _ = system('clear')

def findString(start,sheet, string):
    if type(string) != str: return [0,0]
    string = string.lower()
    for coluna in range(start[0],30): #colunas com nomes 1 e 15
        for linha in range(start[1],320): #linhas com nomes começa com 1 e a cada 20
            #print(coluna,linha)
            celula = (sheet.cell(column=coluna, row=linha).value)
            if type(celula) != str: continue
            celula = celula.lower()
            if string in celula:
                return [coluna,linha]
    return [0, 0] #caso não encontre, retorna 0,0

def preenche_ocupacao(wb2,count,index,qtd_aulas,aulas):

    planilha2 = wb2.sheetnames[0]
    sheet2 = wb2[planilha2]

    diasDaSemana = ('seg1','seg2','ter1','ter2','qua1','qua2','qui1','qui2','sex1','sex2')
    a = 0
    while count > 0:
        for x in range (10):
            docente1 = aulas[index[a]][diasDaSemana[x],'doc1']
            docente2 = aulas[index[a]][diasDaSemana[x],'doc2']
            coord_doc1 = findString([1,1],sheet2,docente1)
            coord_doc2 = findString([1,1],sheet2,docente2)
            #print (a)
            #print (x)
            if coord_doc1 == [0,0] and coord_doc2 == [0,0]: continue
            if coord_doc2 == [0,0]: coord_doc2 = coord_doc1
            if coord_doc1 == [0,0]: coord_doc1 = coord_doc2

            if index[a][0] == 'T':
                coord_doc1[1] += 6
                if coord_doc2 != coord_doc1:
                    coord_doc2[1] += 6
            if index[a][0] == 'N':
               coord_doc1[1] += 12
               if coord_doc2 != coord_doc1:
                    coord_doc2[1] += 12
            for y in range ((qtd_aulas+1)//2):
                if qtd_aulas == 4:
#-------------------------------------------------------------------------------------------------------#
                    if sheet2.cell(column=coord_doc1[0]+2+x-x%2,row=coord_doc1[1]+1+y).value != None:
                        sheet2.cell(column=coord_doc1[0]+2+x-x%2,row=coord_doc1[1]+1+y).value = '-XXX-'
                    else:
                        sheet2.cell(column=coord_doc1[0]+2+x-x%2,row=coord_doc1[1]+1+y).value = aulas[index[a]][diasDaSemana[x],'A'+str(y+1)]

                    if sheet2.cell(column=coord_doc2[0]+2+x-x%2,row=coord_doc2[1]+3+y).value != None:
                        sheet2.cell(column=coord_doc2[0]+2+x-x%2,row=coord_doc2[1]+3+y).value = '-XXX'
                    else:
                        sheet2.cell(column=coord_doc2[0]+2+x-x%2,row=coord_doc2[1]+3+y).value = aulas[index[a]][diasDaSemana[x],'A'+str(y+3)]

                    if aulas[index[a]][diasDaSemana[x],'A'+str(y+1)] != None: sheet2.cell(column=coord_doc1[0]+1+x-x%2,row=coord_doc1[1]+1+y).value = index[a]
                    if aulas[index[a]][diasDaSemana[x],'A'+str(y+3)] != None: sheet2.cell(column=coord_doc2[0]+1+x-x%2,row=coord_doc2[1]+3+y).value = index[a]

#------------------------------------------------------------------------------------------------------#

                else:
                    if sheet2.cell(column=coord_doc1[0]+2+x-x%2,row=coord_doc1[1]+1+y).value != None:
                        sheet2.cell(column=coord_doc1[0]+2+x-x%2,row=coord_doc1[1]+1+y).value = '=XXX='
                    else:
                        sheet2.cell(column=coord_doc1[0]+2+x-x%2,row=coord_doc1[1]+1+y).value = aulas[index[a]][diasDaSemana[x],'A'+str(y+1)]

                    if sheet2.cell(column=coord_doc2[0]+2+x-x%2,row=coord_doc2[1]+4+y).value != None:
                        sheet2.cell(column=coord_doc2[0]+2+x-x%2,row=coord_doc2[1]+4+y).value = '=XXX='
                    else:
                        sheet2.cell(column=coord_doc2[0]+2+x-x%2,row=coord_doc2[1]+4+y).value = aulas[index[a]][diasDaSemana[x],'A'+str(y+4)]

                    if aulas[index[a]][diasDaSemana[x],'A'+str(y+1)] != None: sheet2.cell(column=coord_doc1[0]+1+x-x%2,row=coord_doc1[1]+1+y).value = index[a]
                    if aulas[index[a]][diasDaSemana[x],'A'+str(y+4)] != None: sheet2.cell(column=coord_doc2[0]+1+x-x%2,row=coord_doc2[1]+4+y).value = index[a]
        count -=1
        a +=1
    return wb2



def main():
    print ('Insira o nome do arquivo do Excel que contém o horário escolar:')
    arquivo = input()+'.xlsx'
    qtd_aulas = 0
    clearScreen()
    try:
        wb1 = openpyxl.load_workbook(arquivo)
    except:
        print ('Erro ao abrir o arquivo. O nome está correto?')
        return
    try:
        wb2 = openpyxl.load_workbook('GERAL.xlsx')
    except:
        print ('Erro. O arquivo GERAL.xlsx está na mesma pasta do programa?')

    horario = 3#len(wb1.sheetnames)
    for abas in range (horario):
        planilha1 = wb1.sheetnames[abas]
        sheet1 = wb1[planilha1]
        qtd_aulas = int (sheet1.cell(column=4,row=2).value)# número da linha que informa a quantidade de aulas (na célula'D2')
        linha = 3
        count = 0
        turma = {}
        aulas = {}
        index = {}
    # Cria uma biblioteca aulas, para cada turma do horário
        while (sheet1.cell(column=4,row=linha).value) != 'END': #No final da planilha deve ser colocado END
            turma [count] = Turma
            title,aulas [title]  = turma[count].turma([4,linha],sheet1,qtd_aulas)
            index[count] = title
            if qtd_aulas == 4:
                linha += 13
            else:
                linha +=15
            count += 1

        print(str(abas)+'/'+str(horario)+' Pass!')
    # - - - - -

        wb2 = preenche_ocupacao(wb2,count,index,qtd_aulas,aulas)
    try:
        wb2.save('Ocupacao_docentes.xlsx')
        print(str(horario)+'/'+str(horario)+' Feito!')
    except:
        print('O arquivo Ocupacao_docentes está aberto. Feche-o antes de usar esse programa.')
        return










if __name__ == '__main__':
    main()
